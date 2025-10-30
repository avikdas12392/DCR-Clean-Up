#!/usr/bin/env python3
import warnings; warnings.filterwarnings("ignore", message="The global interpreter lock")
import json, time, re, math, sqlite3
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional

import requests
from openpyxl import load_workbook, Workbook

# ---------- Optional fuzzy lib; fallback to difflib ----------
try:
    from rapidfuzz import fuzz
    def token_set_ratio(a: str, b: str) -> float:
        return float(fuzz.token_set_ratio(a or "", b or ""))
    def partial_ratio(a: str, b: str) -> float:
        return float(fuzz.partial_ratio(a or "", b or ""))
except Exception:
    from difflib import SequenceMatcher
    def _ratio(a: str, b: str) -> float:
        return 100.0 * SequenceMatcher(None, (a or ""), (b or "")).ratio()
    token_set_ratio = _ratio
    partial_ratio    = _ratio

# ================= CONFIG =================
API_KEY = "2e823b299673ed947dde2960a6b609da1f104d15"  # Serper API
ENDPOINT = "https://google.serper.dev/maps"

INPUT_XLSX   = Path("input.xlsx")         # must contain the columns below
OUTPUT1_XLSX = Path("output 1.xlsx")      # allowed results only
OUTPUT2_XLSX = Path("output 2.xlsx")      # one row per input, best match (blank if none)
ELIM_XLSX    = Path("eliminated.xlsx")    # non-allowed categories for audit
PROGRESS_LOG = Path("progress_log.json")

# Input schema (your current headers)
INPUT_COLS = [
    "HQ Name","Doctor Contact ID","UIN No","Customer Name",
    "Associate Hospital","OLong","OLat","Tagged Address","pin"
]

# Output-1 schema (append rating/reviews)
OUTPUT1_COLS = [
    "Name","Address","Lat","Long","Website","Pincode","Category","CID",
    "InputRowIndex","Rating","Reviews"
]

# Output-2 schema = all input cols + selected result cols + score fields
OUTPUT2_COLS = INPUT_COLS + [
    "Match_Name","Match_Address","Match_Lat","Match_Long","Match_Pincode",
    "Match_Category","Match_CID","Match_Rating","Match_Reviews","Match_Score"
]

ELIM_COLS = ["Reason"] + OUTPUT1_COLS  # reason explains why it was filtered

# Search behaviour
NUM_RESULTS     = 20
RADIUS_METERS   = 100     # per your latest instruction
TIMEOUT         = 30
BASE_SLEEP      = 0.8
BACKOFF_FACTOR  = 1.6
MAX_RETRIES     = 4

# Category whitelist (case-insensitive). "Hospital department" REMOVED.
ALLOWED_CATS = {
    "hospital","nursing home","private hospital","general hospital",
    "government hospital","specialized hospital","heart hospital"
}

# Preferred category ordering for tie-breaks (index smaller = better)
CAT_PREFERENCE = [
    "hospital","general hospital","private hospital","government hospital",
    "specialized hospital","heart hospital","nursing home"
]

# Vicinity cache and global de-dupe (to save credits + avoid duplicates)
CACHE_DB_FILE  = Path("vicinity_cache.sqlite3")
DEDUP_DB_FILE  = Path("serper_seen.sqlite3")

# ================ Utilities =================
def now_iso() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def load_progress() -> Dict[str, Any]:
    if not PROGRESS_LOG.exists():
        return {"last_processed_input_row": -1, "errors": []}
    try:
        return json.loads(PROGRESS_LOG.read_text(encoding="utf-8"))
    except Exception as e:
        PROGRESS_LOG.rename(PROGRESS_LOG.with_suffix(".corrupt.json"))
        return {"last_processed_input_row": -1, "errors": [{"row": -1, "error": f"Corrupt progress: {e}", "time": now_iso()}]}

def save_progress(progress: Dict[str, Any]) -> None:
    tmp = PROGRESS_LOG.with_suffix(".tmp")
    tmp.write_text(json.dumps(progress, indent=2), encoding="utf-8")
    tmp.replace(PROGRESS_LOG)

def log_error(progress: Dict[str, Any], row_idx: int, err: str) -> None:
    progress.setdefault("errors", []).append({"row": row_idx, "error": err, "time": now_iso()})
    save_progress(progress)

def ensure_excel_with_sheet(path: Path, columns: List[str], sheet_name: str = "Sheet1") -> None:
    if path.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(columns)
    wb.save(path)

def read_input_rows_xlsx(path: Path, sheet_name: str = "Sheet1") -> List[Dict[str, str]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h or "").strip() for h in rows[0]]
    out: List[Dict[str, str]] = []
    for r in rows[1:]:
        d = {headers[j]: ("" if r is None or j >= len(r) or r[j] is None else str(r[j])) for j in range(len(headers))}
        out.append(d)
    return out

def append_rows_excel(path: Path, rows: List[Dict[str, Any]], columns: List[str], sheet_name: str = "Sheet1") -> None:
    if not rows:
        return
    if not path.exists():
        ensure_excel_with_sheet(path, columns, sheet_name)
    wb = load_workbook(path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    if ws.max_row == 0:
        ws.append(columns)
    for r in rows:
        ws.append([r.get(c, "") for c in columns])
    wb.save(path)

def to_float(x: Any) -> Optional[float]:
    try:
        return float(x)
    except Exception:
        return None

def km_distance(lat1, lon1, lat2, lon2) -> Optional[float]:
    if None in (lat1, lon1, lat2, lon2):
        return None
    R = 6371.0088
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dl   = math.radians(lon2 - lon1)
    a = math.sin(dphi/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return 2*R*math.asin(math.sqrt(a))

def extract_pincode(address: str) -> str:
    if not address:
        return ""
    m = re.search(r"\b(\d{6})\b", address)
    if m:
        return m.group(1)
    parts = [p.strip() for p in address.split(",") if p.strip()]
    if len(parts) >= 2:
        m2 = re.search(r"\b(\d{6})\b", parts[-2])
        if m2:
            return m2.group(1)
    return ""

def norm_cat(c: str) -> str:
    return (c or "").strip().lower()

def is_allowed_category(category: str) -> bool:
    c = norm_cat(category)
    if not c:
        return False
    # exact match or contains (e.g., “General hospital | XYZ”)
    return any(a == c or a in c for a in ALLOWED_CATS)

def cat_rank(category: str) -> int:
    c = norm_cat(category)
    for i, name in enumerate(CAT_PREFERENCE):
        if name == c or name in c:
            return i
    return 999

def brand_tokens(s: str) -> set:
    # crude brand cue extractor — keeps alpha tokens >=4 chars
    toks = re.findall(r"[A-Za-z]{4,}", (s or "").lower())
    # collapse "manipal hospitals" -> {"manipal","hospital(s)"}; ignore generic tokens
    stop = {"hospital","hospitals","general","private","specialized","government","nursing","clinic","department"}
    return {t for t in toks if t not in stop}

def stable_place_key(p: Dict[str, Any]) -> str:
    cid = (p.get("cid") or "").strip()
    if cid:
        return f"cid:{cid}"
    name = (p.get("title") or "").strip().lower()
    addr = (p.get("address") or "").strip().lower()
    return f"na:{name}|{addr}"

# -------- SQLite: global dedupe --------
def open_dedupe_db() -> sqlite3.Connection:
    db = sqlite3.connect(DEDUP_DB_FILE)
    db.execute("PRAGMA journal_mode=WAL;")
    db.execute("PRAGMA synchronous=NORMAL;")
    db.execute("""CREATE TABLE IF NOT EXISTS places_seen (
                    key TEXT PRIMARY KEY,
                    first_seen_at INTEGER DEFAULT (strftime('%s','now'))
                  );""")
    db.commit()
    return db

def is_new_globally_and_mark(db: sqlite3.Connection, place_key: str) -> bool:
    try:
        db.execute("INSERT INTO places_seen (key) VALUES (?);", (place_key,))
        db.commit()
        return True
    except sqlite3.IntegrityError:
        return False

# -------- SQLite: vicinity cache --------
def open_cache_db() -> sqlite3.Connection:
    db = sqlite3.connect(CACHE_DB_FILE)
    db.execute("PRAGMA journal_mode=WAL;")
    db.execute("PRAGMA synchronous=NORMAL;")
    db.execute("""CREATE TABLE IF NOT EXISTS vicinity_cache (
                    vkey TEXT PRIMARY KEY,
                    response_json TEXT NOT NULL,
                    last_used INTEGER DEFAULT (strftime('%s','now'))
                  );""")
    db.commit()
    return db

_in_memory_cache: Dict[str, Dict[str, Any]] = {}

def make_vkey(olat: Any, olon: Any, pin: str) -> str:
    def rd(x, n=5):
        try:
            return f"{round(float(x), n)}"
        except Exception:
            return "NA"
    return f"{pin}|{rd(olat)}|{rd(olon)}|r{RADIUS_METERS}|n{NUM_RESULTS}"

def cache_get(db: sqlite3.Connection, vkey: str) -> Optional[Dict[str, Any]]:
    js = _in_memory_cache.get(vkey)
    if js is not None:
        return js
    cur = db.execute("SELECT response_json FROM vicinity_cache WHERE vkey=?;", (vkey,))
    row = cur.fetchone()
    if row:
        try:
            js = json.loads(row[0])
            _in_memory_cache[vkey] = js
            db.execute("UPDATE vicinity_cache SET last_used=strftime('%s','now') WHERE vkey=?;", (vkey,))
            db.commit()
            return js
        except Exception:
            return None
    return None

def cache_set(db: sqlite3.Connection, vkey: str, js: Dict[str, Any]) -> None:
    payload = json.dumps(js, separators=(",", ":"))
    _in_memory_cache[vkey] = js
    db.execute("INSERT OR REPLACE INTO vicinity_cache (vkey, response_json, last_used) VALUES (?, ?, strftime('%s','now'));",
               (vkey, payload))
    db.commit()

# -------- API call (with radius + ll + keyword/category=hospital) --------
def call_serper(cache_db: sqlite3.Connection, keyword: str, olat: Any, olon: Any, pin: str) -> Optional[Dict[str, Any]]:
    vkey = make_vkey(olat, olon, pin)
    js = cache_get(cache_db, vkey)
    if js is not None:
        print(f"[{now_iso()}] Cache hit → {len(js.get('places', []) or [])} places")
        return js

    headers = {"X-API-KEY": API_KEY, "Content-Type": "application/json"}
    q = f"{keyword} near {olat},{olon} {pin}".strip()
    payload = {
        "q": q,
        "num": NUM_RESULTS,
        "start": 0,
        "ll": f"{olat},{olon}",
        "radius": RADIUS_METERS,
        "category": "hospital"
    }

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = requests.post(ENDPOINT, headers=headers, json=payload, timeout=TIMEOUT)
            resp.raise_for_status()
            js = resp.json()
            if not isinstance(js, dict) or "places" not in js:
                raise ValueError("Unexpected response (no 'places').")
            print(f"[{now_iso()}] API OK → {len(js.get('places', []) or [])} places")
            cache_set(cache_db, vkey, js)
            return js
        except Exception as e:
            if attempt == MAX_RETRIES:
                print(f"[{now_iso()}] API failed after {MAX_RETRIES} attempts: {e}")
                return None
            sleep_for = BASE_SLEEP * (BACKOFF_FACTOR ** (attempt - 1))
            print(f"[{now_iso()}] Retry {attempt}/{MAX_RETRIES} in {sleep_for:.1f}s … ({e})")
            time.sleep(sleep_for)
    return None

# -------- Matching logic (multi-signal ranker) --------
def score_candidate(inp: Dict[str, str], cand: Dict[str, Any]) -> float:
    # Pull inputs
    assoc = (inp.get("Associate Hospital") or "").lower()
    tag_addr = (inp.get("Tagged Address") or "").lower()
    pin_in   = (inp.get("pin") or "").strip()

    olat = to_float(inp.get("OLat"))
    olon = to_float(inp.get("OLong"))

    name = (cand.get("title") or "")
    addr = (cand.get("address") or "")
    lat2 = to_float(cand.get("latitude"))
    lon2 = to_float(cand.get("longitude"))
    cat  = (cand.get("category") or cand.get("type") or "")

    # Signals
    # 1) Distance (0..40)
    dist_km = km_distance(olat, olon, lat2, lon2)
    dist_m  = 1000.0 * dist_km if dist_km is not None else None
    if dist_m is None:
        dist_score = 0.0
    elif dist_m <= 100:
        # full points within 30m; linear fade to 0 at 150m
        dist_score = 40.0 * max(0.0, min(1.0, (150.0 - max(30.0, dist_m)) / 120.0))
    else:
        dist_score = 0.0

    # 2) Pin bonus (0..10)
    pin_out = extract_pincode(addr)
    pin_bonus = 10.0 if pin_in and pin_out and pin_in == pin_out else 0.0

    # 3) Name similarity vs Associate Hospital (0..20)
    name_sim = token_set_ratio(assoc, name.lower())  # 0..100
    name_score = 0.20 * name_sim  # scale to 20

    # 4) Address similarity vs Tagged Address (0..20)
    addr_sim = token_set_ratio(tag_addr, addr.lower())
    addr_score = 0.20 * addr_sim

    # 5) Brand cue (0..6)
    b_in  = brand_tokens(assoc + " " + tag_addr)
    b_out = brand_tokens(name + " " + addr)
    brand_overlap = len(b_in & b_out)
    brand_bonus = 6.0 if brand_overlap >= 1 else 0.0

    # 6) Category boost (0..4)
    cat_boost = 4.0 if is_allowed_category(cat) else 0.0

    # Total (max ~100)
    total = dist_score + pin_bonus + name_score + addr_score + brand_bonus + cat_boost
    return round(total, 2)

def pick_best_candidate(inp: Dict[str, str], places: List[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    if not places:
        return None

    # Filter to allowed categories first
    allowed = [p for p in places if is_allowed_category(p.get("category") or p.get("type") or "")]
    if not allowed:
        return None

    # Score & sort
    scored = []
    for p in allowed:
        s = score_candidate(inp, p)
        rated = to_float(p.get("rating")) or 0.0
        reviews = int(p.get("ratingCount") or 0)
        scored.append((s, -reviews, cat_rank(p.get("category") or p.get("type") or ""), -rated, p))

    scored.sort(reverse=True, key=lambda x: (x[0], x[1], -x[2], x[3]))
    best = scored[0][-1] if scored else None
    return best

# ================= MAIN =================
def main():
    # Ensure outputs exist with headers
    ensure_excel_with_sheet(OUTPUT1_XLSX, OUTPUT1_COLS)
    ensure_excel_with_sheet(OUTPUT2_XLSX, OUTPUT2_COLS)
    ensure_excel_with_sheet(ELIM_XLSX,    ELIM_COLS)

    progress = load_progress()
    last_done = int(progress.get("last_processed_input_row", -1))

    # Read input rows
    try:
        input_rows = read_input_rows_xlsx(INPUT_XLSX)
    except Exception as e:
        print(f"[{now_iso()}] FATAL: cannot read input: {e}")
        log_error(progress, -1, f"Input read error: {e}")
        return

    # Validate columns (soft check)
    missing = [c for c in INPUT_COLS if c not in (input_rows and input_rows[0].keys())]
    if missing:
        print(f"[{now_iso()}] WARNING: Missing expected columns: {missing}")

    dedupe_db = open_dedupe_db()
    cache_db  = open_cache_db()

    n = len(input_rows)
    print(f"[{now_iso()}] Loaded input with {n} rows. Resuming from row {last_done + 1}.")

    for idx in range(last_done + 1, n):
        row = input_rows[idx]
        try:
            assoc  = row.get("Associate Hospital","")
            olon   = row.get("OLong","")
            olat   = row.get("OLat","")
            pin    = row.get("pin","")

            print(f"[{now_iso()}] Row {idx}: Search 'hospital' @ (olat={olat}, olong={olon}, pin={pin}), r={RADIUS_METERS}m")

            js = call_serper(cache_db, "hospital", olat, olon, pin)
            places = (js or {}).get("places", []) or []

            # Per-row write buffers
            out1_rows: List[Dict[str, Any]] = []
            elim_rows: List[Dict[str, Any]] = []

            # Stream results to Output-1 / Eliminated
            seen_row_keys = set()
            for p in places[:NUM_RESULTS]:
                k = stable_place_key(p)
                if k in seen_row_keys:
                    continue
                seen_row_keys.add(k)

                name = p.get("title","")
                addr = p.get("address","")
                lat2 = p.get("latitude","")
                lon2 = p.get("longitude","")
                web  = p.get("website","") or ""
                cat  = (p.get("category") or p.get("type") or "")
                cid  = (p.get("cid") or "").strip()
                pin_out = extract_pincode(addr)
                rating  = p.get("rating","")
                reviews = p.get("ratingCount","")

                r = {
                    "Name": name,
                    "Address": addr,
                    "Lat": lat2, "Long": lon2,
                    "Website": web, "Pincode": pin_out,
                    "Category": cat, "CID": cid,
                    "InputRowIndex": idx,
                    "Rating": rating, "Reviews": reviews
                }

                if is_allowed_category(cat):
                    # global dedupe for Output-1
                    if is_new_globally_and_mark(dedupe_db, k):
                        out1_rows.append(r)
                else:
                    elim_rows.append({"Reason":"Category not allowed", **r})

            if out1_rows:
                append_rows_excel(OUTPUT1_XLSX, out1_rows, OUTPUT1_COLS)
            if elim_rows:
                append_rows_excel(ELIM_XLSX, elim_rows, ELIM_COLS)

            # Pick BEST candidate for Output-2 (may be None)
            best = pick_best_candidate(row, places)

            if best:
                out2 = {
                    **{c: row.get(c, "") for c in INPUT_COLS},
                    "Match_Name": best.get("title",""),
                    "Match_Address": best.get("address",""),
                    "Match_Lat": best.get("latitude",""),
                    "Match_Long": best.get("longitude",""),
                    "Match_Pincode": extract_pincode(best.get("address","")),
                    "Match_Category": (best.get("category") or best.get("type") or ""),
                    "Match_CID": (best.get("cid") or "").strip(),
                    "Match_Rating": best.get("rating",""),
                    "Match_Reviews": best.get("ratingCount",""),
                    "Match_Score": score_candidate(row, best)
                }
            else:
                # No allowed candidate — still write the input row with blanks
                out2 = {c: row.get(c, "") for c in INPUT_COLS}
                out2.update({
                    "Match_Name":"","Match_Address":"","Match_Lat":"","Match_Long":"",
                    "Match_Pincode":"","Match_Category":"","Match_CID":"",
                    "Match_Rating":"","Match_Reviews":"","Match_Score":""
                })

            append_rows_excel(OUTPUT2_XLSX, [out2], OUTPUT2_COLS)

            # Save progress
            progress["last_processed_input_row"] = idx
            save_progress(progress)
            print(f"[{now_iso()}]   ✓ Row {idx} done. Output1+Elim wrote {len(out1_rows)}+{len(elim_rows)}")

        except Exception as e:
            print(f"[{now_iso()}]   ✗ Row {idx} error: {e}")
            log_error(progress, idx, str(e))
            progress["last_processed_input_row"] = idx
            save_progress(progress)
            continue

    print(f"[{now_iso()}] All done. Progress saved: {PROGRESS_LOG.name}")

if __name__ == "__main__":
    main()
