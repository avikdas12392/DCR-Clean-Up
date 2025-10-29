#!/usr/bin/env python3
import warnings; warnings.filterwarnings("ignore", message="The global interpreter lock")

import json
import time
import re
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple

import requests
from openpyxl import load_workbook, Workbook

# ---------- Optional fuzzy lib; fallback to difflib ----------
try:
    from rapidfuzz.fuzz import ratio as fuzz_ratio
    def fuzzy_score(a: str, b: str) -> float:
        return float(fuzz_ratio(a or "", b or ""))
except Exception:
    from difflib import SequenceMatcher
    def fuzzy_score(a: str, b: str) -> float:
        return 100.0 * SequenceMatcher(None, (a or ""), (b or "")).ratio()

# ================= CONFIG =================
API_KEY = "2e823b299673ed947dde2960a6b609da1f104d15"  # Serper API
ENDPOINT = "https://google.serper.dev/maps"

INPUT_FILE = Path("input.xlsx")
OUTPUT1_FILE = Path("output 1.xlsx")   # Name, Address, Lat, Long, Website, Pincode, Category, CID, InputRowIndex, Rating, Reviews
OUTPUT2_FILE = Path("output 2.xlsx")   # [all input cols] + [output1 cols] + FuzzyScore
PROGRESS_LOG = Path("progress_log.json")

# Per-call behavior
NUM_RESULTS = 20
RADIUS_METERS = 200          # Set 0 to omit radius entirely
TIMEOUT = 30
BASE_SLEEP = 0.8
BACKOFF_FACTOR = 1.6
MAX_RETRIES = 4

FUZZY_THRESHOLD = 75.0
VICINITY_DECIMALS = 3         # 3 ≈ ~100m, 4 ≈ ~10m

# India guardrails
INDIA_LAT_MIN, INDIA_LAT_MAX = 6.0, 37.5
INDIA_LON_MIN, INDIA_LON_MAX = 68.0, 97.5

# SQLite files
DEDUP_DB_FILE = Path("serper_seen.sqlite3")      # global dedupe across the whole run
CACHE_DB_FILE = Path("vicinity_cache.sqlite3")   # persistent cache per vicinity-key

# Output 1 schema
OUTPUT1_COLS_FIXED = [
    "Name","Address","Lat","Long","Website","Pincode","Category","CID","InputRowIndex","Rating","Reviews"
]

# ================= Utilities =================
def now_iso() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def load_progress() -> Dict[str, Any]:
    if not PROGRESS_LOG.exists():
        return {"last_processed_input_row": -1, "errors": []}
    try:
        return json.loads(PROGRESS_LOG.read_text(encoding="utf-8"))
    except Exception as e:
        PROGRESS_LOG.rename(PROGRESS_LOG.with_suffix(".corrupt.json"))
        return {"last_processed_input_row": -1, "errors": [{"row": -1, "error": f"Corrupt progress: {e}", "time": now_iso()}]}

def save_progress(progress: Dict[str, Any]) -> None:
    tmp = PROGRESS_LOG.with_suffix(".tmp")
    tmp.write_text(json.dumps(progress, indent=2), encoding="utf-8")
    tmp.replace(PROGRESS_LOG)

def log_error(progress: Dict[str, Any], row_idx: int, err: str) -> None:
    progress.setdefault("errors", []).append({"row": row_idx, "error": err, "time": now_iso()})
    save_progress(progress)

def print_call_purpose(row_idx: int, olat: Any, olon: Any, pin: str, assoc: str, using_cache: bool = False):
    src = "Cache" if using_cache else "Serper Maps API"
    print(f"[{now_iso()}] Row {row_idx}: Search '{assoc or 'hospital'} + hospital' within {RADIUS_METERS}m via {src}. Inputs: OLat={olat}, OLong={olon}, pin={pin}. Limit={NUM_RESULTS}")

def extract_pincode(address: str) -> str:
    if not address:
        return ""
    m = re.search(r"\b(\d{6})\b", address)
    if m:
        return m.group(1)
    parts = [p.strip() for p in address.split(",") if p.strip()]
    if len(parts) >= 2:
        candidate = parts[-2]
        m2 = re.search(r"\b(\d{6})\b", candidate)
        if m2:
            return m2.group(1)
    return ""

def stable_place_key(p: Dict[str, Any]) -> str:
    cid = (p.get("cid") or "").strip()
    if cid:
        return f"cid:{cid}"
    name = (p.get("title") or "").strip().lower()
    addr = (p.get("address") or "").strip().lower()
    return f"na:{name}|{addr}"

def float_or_none(x: Any) -> Optional[float]:
    try:
        return float(x)
    except Exception:
        return None

def looks_like_india_latlon(lat: float, lon: float) -> bool:
    try:
        return (INDIA_LAT_MIN <= float(lat) <= INDIA_LAT_MAX) and (INDIA_LON_MIN <= float(lon) <= INDIA_LON_MAX)
    except Exception:
        return False

def in_india_or_has_indian_pin(addr: str, lat: Any, lon: Any) -> bool:
    try:
        latf = float(lat)
        lonf = float(lon)
        if looks_like_india_latlon(latf, lonf):
            return True
    except Exception:
        pass
    addr_l = (addr or "").lower()
    if " india" in addr_l or re.search(r"\b\d{6}\b", addr_l):
        return True
    return False

# -------- openpyxl I/O (no pandas) --------
def ensure_excel_with_sheet(path: Path, columns: List[str], sheet_name: str = "Sheet1") -> None:
    if path.exists():
        wb = load_workbook(path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        if ws.max_row == 0:
            ws.append(columns)
            wb.save(path)
        return
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(columns)
    wb.save(path)

def read_input_rows_xlsx(path: Path, sheet_name: str = "Sheet1") -> Tuple[List[str], List[Dict[str, str]]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(h or "").strip() for h in rows[0]]
    out: List[Dict[str, str]] = []
    for r in rows[1:]:
        d = {headers[j]: ("" if r is None or j >= len(r) or r[j] is None else str(r[j])) for j in range(len(headers))}
        out.append(d)
    return headers, out

def append_rows_excel(path: Path, rows: List[Dict[str, Any]], columns: List[str], sheet_name: str = "Sheet1") -> None:
    if not rows:
        return
    ensure_excel_with_sheet(path, columns, sheet_name)
    wb = load_workbook(path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    for r in rows:
        ws.append([r.get(c, "") for c in columns])
    wb.save(path)

# -------- SQLite: global dedupe --------
def open_dedupe_db() -> sqlite3.Connection:
    db = sqlite3.connect(DEDUP_DB_FILE)
    db.execute("PRAGMA journal_mode=WAL;")
    db.execute("PRAGMA synchronous=NORMAL;")
    db.execute("""CREATE TABLE IF NOT EXISTS places_seen (
                    key TEXT PRIMARY KEY,
                    first_seen_at INTEGER DEFAULT (strftime('%s','now'))
                  );""")
    db.commit()
    return db

def is_new_globally_and_mark(db: sqlite3.Connection, place_key: str) -> bool:
    try:
        db.execute("INSERT INTO places_seen (key) VALUES (?);", (place_key,))
        db.commit()
        return True
    except sqlite3.IntegrityError:
        return False

# -------- SQLite: vicinity cache --------
def open_cache_db() -> sqlite3.Connection:
    db = sqlite3.connect(CACHE_DB_FILE)
    db.execute("PRAGMA journal_mode=WAL;")
    db.execute("PRAGMA synchronous=NORMAL;")
    db.execute("""CREATE TABLE IF NOT EXISTS vicinity_cache (
                    vkey TEXT PRIMARY KEY,
                    response_json TEXT NOT NULL,
                    last_used INTEGER DEFAULT (strftime('%s','now'))
                  );""")
    db.commit()
    return db

_in_memory_cache: Dict[str, Dict[str, Any]] = {}

def make_vicinity_key(olat: Any, olon: Any, pin: str) -> str:
    def rd(x):
        try:
            return f"{round(float(x), VICINITY_DECIMALS)}"
        except Exception:
            return "NA"
    return f"{pin}|{rd(olat)}|{rd(olon)}|r{RADIUS_METERS}|n{NUM_RESULTS}"

def cache_get(db: sqlite3.Connection, vkey: str) -> Optional[Dict[str, Any]]:
    js = _in_memory_cache.get(vkey)
    if js is not None:
        return js
    cur = db.execute("SELECT response_json FROM vicinity_cache WHERE vkey=?;", (vkey,))
    row = cur.fetchone()
    if row:
        try:
            js = json.loads(row[0])
            _in_memory_cache[vkey] = js
            db.execute("UPDATE vicinity_cache SET last_used=strftime('%s','now') WHERE vkey=?;", (vkey,))
            db.commit()
            return js
        except Exception:
            return None
    return None

def cache_set(db: sqlite3.Connection, vkey: str, js: Dict[str, Any]) -> None:
    payload = json.dumps(js, separators=(",", ":"))
    _in_memory_cache[vkey] = js
    db.execute("INSERT OR REPLACE INTO vicinity_cache (vkey, response_json, last_used) VALUES (?, ?, strftime('%s','now'));",
               (vkey, payload))
    db.commit()

# -------- Serper call w/ cache --------
def call_serper_or_cache(cache_db: sqlite3.Connection, assoc: str, olat: Any, olon: Any, pin: str) -> Optional[Dict[str, Any]]:
    vkey = make_vicinity_key(olat, olon, pin)

    # Try cache first
    js = cache_get(cache_db, vkey)
    if js is not None:
        print_call_purpose(-1, olat, olon, pin, assoc, using_cache=True)
        print(f"[{now_iso()}]   ✅ Cache hit: {len(js.get('places', []) or [])} places.")
        return js

    headers = {"X-API-KEY": API_KEY, "Content-Type": "application/json"}

    # >>> ONLY CHANGE REQUESTED: always include the keyword 'hospital' <<<
    base = (assoc or "").strip()
    if base:
        query_term = f"{base} hospital"   # enforce hospital category/keyword
    else:
        query_term = "hospital"

    q = f"{query_term} near {olat},{olon} {pin} India".strip()

    payload = {
        "q": q,
        "num": NUM_RESULTS,
        "start": 0,
        "ll": f"{olat},{olon}",
        "gl": "in",
        "hl": "en"
    }
    if RADIUS_METERS and RADIUS_METERS > 0:
        payload["radius"] = RADIUS_METERS

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = requests.post(ENDPOINT, headers=headers, json=payload, timeout=TIMEOUT)
            resp.raise_for_status()
            js = resp.json()
            if not isinstance(js, dict) or "places" not in js:
                raise ValueError("Unexpected response structure (no 'places').")
            print(f"[{now_iso()}]   ✅ API OK: received {len(js.get('places', []) or [])} places.")
            cache_set(cache_db, vkey, js)
            return js
        except Exception as e:
            if attempt == MAX_RETRIES:
                print(f"[{now_iso()}]   ❌ API failed after {MAX_RETRIES} attempts: {e}")
                return None
            sleep_for = BASE_SLEEP * (BACKOFF_FACTOR ** (attempt - 1))
            print(f"[{now_iso()}]   ~ Retry {attempt}/{MAX_RETRIES} in {sleep_for:.1f}s due to: {e}")
            time.sleep(sleep_for)
    return None

# ================= MAIN =================
def main():
    progress = load_progress()
    last_done = int(progress.get("last_processed_input_row", -1))

    # Read input headers + rows (as list[dict[str,str]])
    try:
        input_headers, input_rows = read_input_rows_xlsx(INPUT_FILE)
        if not input_rows:
            print(f"[{now_iso()}] FATAL: input is empty or unreadable.")
            return
    except Exception as e:
        print(f"[{now_iso()}] FATAL: cannot read input: {e}")
        log_error(progress, -1, f"Input read error: {e}")
        return

    # Validate required columns
    required = ["Associate Hospital", "OLat", "OLong", "pin", "Tagged Address"]
    missing = [c for c in required if c not in input_headers]
    if missing:
        print(f"[{now_iso()}] FATAL: Missing required columns in input: {missing}")
        return

    # Build Output 2 schema dynamically: all input columns + Output1 cols + FuzzyScore
    output1_cols = OUTPUT1_COLS_FIXED[:]
    output2_cols = input_headers[:] + output1_cols + ["FuzzyScore"]

    # Ensure outputs exist with headers (idempotent)
    ensure_excel_with_sheet(OUTPUT1_FILE, output1_cols)
    ensure_excel_with_sheet(OUTPUT2_FILE, output2_cols)

    dedupe_db = open_dedupe_db()
    cache_db = open_cache_db()

    n = len(input_rows)
    print(f"[{now_iso()}] Loaded input with {n} rows. Resuming from row {last_done + 1}.")

    for idx in range(last_done + 1, n):
        row = input_rows[idx]
        try:
            assoc = (row.get("Associate Hospital") or "").strip()
            olat = row.get("OLat", "")
            olon = row.get("OLong", "")
            pin = row.get("pin", "")
            tagged_addr = row.get("Tagged Address", "")

            olat_f = float_or_none(olat)
            olon_f = float_or_none(olon)
            if olat_f is None or olon_f is None:
                log_error(progress, idx, f"Invalid OLat/OLong: OLat={olat}, OLong={olon}")
                progress["last_processed_input_row"] = idx
                save_progress(progress)
                continue

            print_call_purpose(idx, olat_f, olon_f, pin, assoc, using_cache=False)
            js = call_serper_or_cache(cache_db, assoc, olat_f, olon_f, pin)
            if js is None:
                log_error(progress, idx, "API call failed after retries")
                progress["last_processed_input_row"] = idx
                save_progress(progress)
                continue

            places = js.get("places", []) or []

            # per-row dedupe + global dedupe
            seen_row = set()
            out1_rows: List[Dict[str, Any]] = []
            out2_rows: List[Dict[str, Any]] = []

            for p in places[:NUM_RESULTS]:
                key = stable_place_key(p)
                if key in seen_row:
                    continue

                name = p.get("title", "")
                addr = p.get("address", "")
                lat2 = p.get("latitude", "")
                lon2 = p.get("longitude", "")
                website = p.get("website", "") or ""
                cid = (p.get("cid") or "").strip()
                category = p.get("category", "") or p.get("type", "")
                rating = p.get("rating", "")
                reviews = p.get("ratingCount", "")

                if not in_india_or_has_indian_pin(addr, lat2, lon2):
                    continue

                pinc = extract_pincode(addr)

                r1 = {
                    "Name": name,
                    "Address": addr,
                    "Lat": lat2 if lat2 != "" else "",
                    "Long": lon2 if lon2 != "" else "",
                    "Website": website,
                    "Pincode": pinc,
                    "Category": category,
                    "CID": cid,
                    "InputRowIndex": idx,
                    "Rating": rating,
                    "Reviews": reviews,
                }

                seen_row.add(key)

                if is_new_globally_and_mark(dedupe_db, key):
                    out1_rows.append(r1)

                score = fuzzy_score((addr or "").lower(), (tagged_addr or "").lower())
                if score >= FUZZY_THRESHOLD:
                    left = {c: row.get(c, "") for c in input_headers}
                    out2_rows.append({**left, **r1, "FuzzyScore": round(score, 2)})

            if out1_rows:
                append_rows_excel(OUTPUT1_FILE, out1_rows, output1_cols)
            if out2_rows:
                append_rows_excel(OUTPUT2_FILE, out2_rows, output2_cols)

            print(f"[{now_iso()}]   ✅ Wrote {len(out1_rows)} new uniques to 'output 1', {len(out2_rows)} matches to 'output 2'.")

            progress["last_processed_input_row"] = idx
            save_progress(progress)

        except Exception as e:
            print(f"[{now_iso()}]   ❌ Row {idx} error: {e}")
            log_error(progress, idx, str(e))
            progress["last_processed_input_row"] = idx
            save_progress(progress)
            continue

    print(f"[{now_iso()}] All done. Progress saved to {PROGRESS_LOG.name}.")


if __name__ == "__main__":
    main()
