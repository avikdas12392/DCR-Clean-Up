#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Long-run hospital matcher:
- Reads input.xlsx rows (Associate Hospital, OLat, OLong, pin, Tagged Address, etc.)
- Calls Serper Maps (q = "<Associate Hospital> hospital near <OLat,OLong> <pin> India")
- Whitelists hospital-like categories; non-whitelisted → 'Eliminated' sheet
- Global dedupe + vicinity cache to save credits
- Robust, resumable via progress_log.json
- Output 1: allowed + eliminated (with Rating, Reviews)
- Output 2: only whitelisted; picks ONE best candidate per input row using blended score
"""

import warnings; warnings.filterwarnings("ignore", message="The global interpreter lock")

import json
import time
import re
import math
import sqlite3
import unicodedata
import string
from functools import lru_cache
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple

import requests
from openpyxl import load_workbook, Workbook

# ---------- Optional fuzzy lib; fallback to difflib ----------
try:
    from rapidfuzz.fuzz import ratio as fuzz_ratio
    def fuzzy_score(a: str, b: str) -> float:
        return float(fuzz_ratio(a or "", b or ""))
except Exception:
    from difflib import SequenceMatcher
    def fuzzy_score(a: str, b: str) -> float:
        return 100.0 * SequenceMatcher(None, (a or ""), (b or "")).ratio()

# ================= CONFIG =================
API_KEY = "2e823b299673ed947dde2960a6b609da1f104d15"  # Serper API key
ENDPOINT = "https://google.serper.dev/maps"

INPUT_FILE = Path("input.xlsx")
OUTPUT1_FILE = Path("output 1.xlsx")   # Sheet1 = allowed; Eliminated = filtered-out
OUTPUT2_FILE = Path("output 2.xlsx")
PROGRESS_LOG = Path("progress_log.json")

NUM_RESULTS = 20
RADIUS_METERS = 200              # can be 0 to omit; we still rank with distance
TIMEOUT = 30
BASE_SLEEP = 0.8
BACKOFF_FACTOR = 1.6
MAX_RETRIES = 4
VICINITY_DECIMALS = 3            # 3~100m; 4~10m cache granularity

# Match thresholds / funnel sizes
SHORTLIST_NEAREST = 5            # shortlist top-N by distance for scoring
OUT2_MIN_SCORE = 75.0            # require this score to write to Output 2

# Simple India & PIN guards
INDIA_LAT_MIN, INDIA_LAT_MAX = 6.0, 37.5
INDIA_LON_MIN, INDIA_LON_MAX = 68.0, 97.5

# Files for persistent ledgers
DEDUP_DB_FILE = Path("serper_seen.sqlite3")      # global Output1 dedupe
CACHE_DB_FILE = Path("vicinity_cache.sqlite3")   # persistent response cache

# Output columns
OUTPUT1_COLS = [
    "Name","Address","Lat","Long","Website","Pincode","Category","CID",
    "InputRowIndex","Rating","Reviews"
]
# Output 2 = all input columns (preserved order) + OUTPUT1_COLS + FuzzyScore
INPUT_REQUIRED = ["Associate Hospital","OLat","OLong","pin","Tagged Address"]

# ✅ Category whitelist (case-insensitive). Only these go to Output 1 (Sheet1) & Output 2
WHITELIST_CATEGORIES = {
    "hospital",
    "nursing home",
    "private hospital",
    "general hospital",
    "government hospital",
    "specialized hospital",
    "heart hospital",
    "hospital department",
}

# ================= Utilities =================
def now_iso() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def load_progress() -> Dict[str, Any]:
    if not PROGRESS_LOG.exists():
        return {"last_processed_input_row": -1, "errors": []}
    try:
        return json.loads(PROGRESS_LOG.read_text(encoding="utf-8"))
    except Exception as e:
        PROGRESS_LOG.rename(PROGRESS_LOG.with_suffix(".corrupt.json"))
        return {"last_processed_input_row": -1,
                "errors":[{"row":-1,"error":f"Corrupt progress:{e}","time":now_iso()}]}

def save_progress(p: Dict[str, Any]) -> None:
    tmp = PROGRESS_LOG.with_suffix(".tmp")
    tmp.write_text(json.dumps(p, indent=2), encoding="utf-8")
    tmp.replace(PROGRESS_LOG)

def log_error(p: Dict[str, Any], row: int, err: str) -> None:
    p.setdefault("errors", []).append({"row":row,"error":err,"time":now_iso()})
    save_progress(p)

def print_call_purpose(i: int, olat: Any, olong: Any, pin: str, assoc: str, using_cache: bool=False):
    src = "Cache" if using_cache else "Serper Maps API"
    print(f"[{now_iso()}] Row {i}: Search '{(assoc or '').strip()} + hospital' via {src}. OLat={olat}, OLong={olong}, pin={pin}, num={NUM_RESULTS}")

def extract_pincode(address: str) -> str:
    if not address:
        return ""
    m = re.search(r"\b(\d{6})\b", address)
    if m:
        return m.group(1)
    parts = [p.strip() for p in address.split(",") if p.strip()]
    if len(parts) >= 2:
        m2 = re.search(r"\b(\d{6})\b", parts[-2])
        if m2:
            return m2.group(1)
    return ""

def stable_place_key(p: Dict[str, Any]) -> str:
    cid = (p.get("cid") or "").strip()
    if cid:
        return f"cid:{cid}"
    return f"na:{(p.get('title','').lower().strip())}|{(p.get('address','').lower().strip())}"

def float_or_none(x: Any) -> Optional[float]:
    try:
        return float(x)
    except Exception:
        return None

def looks_like_india(lat: Any, lon: Any) -> bool:
    try:
        return INDIA_LAT_MIN <= float(lat) <= INDIA_LAT_MAX and INDIA_LON_MIN <= float(lon) <= INDIA_LON_MAX
    except Exception:
        return False

def in_india_or_has_pin(addr: str, lat: Any, lon: Any) -> bool:
    # guard against US/etc. places
    try:
        if looks_like_india(lat, lon):
            return True
    except Exception:
        pass
    a = (addr or "").lower()
    return " india" in a or bool(re.search(r"\b\d{6}\b", a))

# ---- Excel I/O ----
def ensure_excel_with_sheet(path: Path, cols: List[str], sheet: str="Sheet1") -> None:
    if path.exists():
        wb = load_workbook(path)
        if sheet in wb.sheetnames:
            ws = wb[sheet]
            if ws.max_row == 0:
                ws.append(cols)
                wb.save(path)
            return
        ws = wb.create_sheet(title=sheet)
        ws.append(cols)
        wb.save(path)
        return
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(cols)
    wb.save(path)

def read_input_rows_xlsx(path: Path) -> Tuple[List[str], List[Dict[str, str]]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(h or "").strip() for h in rows[0]]
    out: List[Dict[str, str]] = []
    for r in rows[1:]:
        out.append({headers[j]: ("" if r is None or j >= len(r) or r[j] is None else str(r[j])) for j in range(len(headers))})
    return headers, out

def append_rows_excel(path: Path, rows: List[Dict[str, Any]], cols: List[str], sheet: str="Sheet1") -> None:
    if not rows:
        return
    ensure_excel_with_sheet(path, cols, sheet)
    wb = load_workbook(path)
    ws = wb[sheet]
    for r in rows:
        ws.append([r.get(c, "") for c in cols])
    wb.save(path)

# ---- SQLite helpers ----
def open_db(file: Path, create_sql: str) -> sqlite3.Connection:
    db = sqlite3.connect(file)
    db.execute("PRAGMA journal_mode=WAL;")
    db.execute("PRAGMA synchronous=NORMAL;")
    db.execute(create_sql)
    db.commit()
    return db

def open_dedupe_db() -> sqlite3.Connection:
    return open_db(
        DEDUP_DB_FILE,
        "CREATE TABLE IF NOT EXISTS places_seen(key TEXT PRIMARY KEY, first_seen_at INTEGER DEFAULT (strftime('%s','now')));"
    )

def is_new_and_mark(db: sqlite3.Connection, key: str) -> bool:
    try:
        db.execute("INSERT INTO places_seen VALUES (?,strftime('%s','now'));", (key,))
        db.commit()
        return True
    except sqlite3.IntegrityError:
        return False

def open_cache_db() -> sqlite3.Connection:
    return open_db(
        CACHE_DB_FILE,
        "CREATE TABLE IF NOT EXISTS vicinity_cache(vkey TEXT PRIMARY KEY,response_json TEXT,last_used INTEGER DEFAULT (strftime('%s','now')));"
    )

_in_mem_cache: Dict[str, Dict[str, Any]] = {}

def make_vkey(olat: Any, olong: Any, pin: str) -> str:
    def rd(x):
        try:
            return f"{round(float(x), VICINITY_DECIMALS)}"
        except Exception:
            return "NA"
    return f"{pin}|{rd(olat)}|{rd(olong)}|n{NUM_RESULTS}|r{RADIUS_METERS}"

def cache_get(db: sqlite3.Connection, vkey: str) -> Optional[Dict[str, Any]]:
    js = _in_mem_cache.get(vkey)
    if js:
        return js
    cur = db.execute("SELECT response_json FROM vicinity_cache WHERE vkey=?;", (vkey,))
    row = cur.fetchone()
    if row:
        try:
            js = json.loads(row[0])
            _in_mem_cache[vkey] = js
            db.execute("UPDATE vicinity_cache SET last_used=strftime('%s','now') WHERE vkey=?;", (vkey,))
            db.commit()
            return js
        except Exception:
            return None
    return None

def cache_set(db: sqlite3.Connection, vkey: str, js: Dict[str, Any]) -> None:
    _in_mem_cache[vkey] = js
    db.execute("INSERT OR REPLACE INTO vicinity_cache VALUES(?, ?, strftime('%s','now'));",
               (vkey, json.dumps(js, separators=(',',':'))))
    db.commit()

# ---- Serper call ----
def call_serper_or_cache(cache_db: sqlite3.Connection, assoc: str, olat: float, olong: float, pin: str) -> Optional[Dict[str, Any]]:
    vkey = make_vkey(olat, olong, pin)
    js = cache_get(cache_db, vkey)
    if js is not None:
        print_call_purpose(-1, olat, olong, pin, assoc, using_cache=True)
        print(f"[{now_iso()}]   ✅ Cache hit: {len(js.get('places', []) or [])} places.")
        return js

    headers = {"X-API-KEY": API_KEY, "Content-Type": "application/json"}
    # ❗ Include Keyword/Category = 'hospital' in query
    q = f"{(assoc or '').strip()} hospital near {olat},{olong} {pin} India".strip()
    payload = {"q": q, "num": NUM_RESULTS, "start": 0, "ll": f"{olat},{olong}", "gl":"in", "hl":"en"}
    if RADIUS_METERS > 0:
        payload["radius"] = RADIUS_METERS

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = requests.post(ENDPOINT, headers=headers, json=payload, timeout=TIMEOUT)
            resp.raise_for_status()
            js = resp.json()
            if not isinstance(js, dict) or "places" not in js:
                raise ValueError("Unexpected response structure (no 'places').")
            print(f"[{now_iso()}]   ✅ API OK: received {len(js.get('places', []) or [])} places.")
            cache_set(cache_db, vkey, js)
            return js
        except Exception as e:
            if attempt == MAX_RETRIES:
                print(f"[{now_iso()}]   ❌ API failed after {MAX_RETRIES} attempts: {e}")
                return None
            sleep_for = BASE_SLEEP * (BACKOFF_FACTOR ** (attempt - 1))
            print(f"[{now_iso()}]   ~ Retry {attempt}/{MAX_RETRIES} in {sleep_for:.1f}s due to: {e}")
            time.sleep(sleep_for)
    return None

# ---------- Text & geo helpers for matching ----------
_PUNCT = str.maketrans({c: " " for c in (string.punctuation + "’‘“”")})
COMMON_SYNONYMS = {
    # spelling variants / normalizations
    "yeshwantpur": "yeshwanthpur",
    "yeshavantpur": "yeshwanthpur",
    "yeshwanthpur": "yeshwanthpur",
    "malleswaram": "malleshwaram",
    "malleshwaram": "malleshwaram",
    "hebbal": "hebbal",
    "hal": "hal",
    "kodihalli": "kodihalli",
    # brands
    "manipal": "manipal",
    "baptist": "baptist",
    "fortis": "fortis",
    "narayana": "narayana",
    "sparsh": "sparsh",
    "heart": "heart",
}
BRAND_TOKENS = {"manipal", "baptist", "fortis", "narayana", "sparsh"}

@lru_cache(maxsize=100_000)
def norm_text(s: str) -> str:
    s = unicodedata.normalize("NFKC", (s or "")).lower()
    s = s.translate(_PUNCT)
    s = " ".join(COMMON_SYNONYMS.get(tok, tok) for tok in s.split())
    return " ".join(s.split())

def token_set(s: str) -> set:
    return set(norm_text(s).split())

def brand_hits(assoc: str, name: str, addr: str) -> int:
    at = token_set(assoc)
    nn = token_set(name) | token_set(addr)
    return sum(1 for b in (BRAND_TOKENS & at) if b in nn)

def locality_hits(tagged_addr: str, cand_addr: str) -> float:
    # Jaccard on non-brand, 3+ char tokens
    ta = {t for t in token_set(tagged_addr) if len(t) >= 3 and t not in BRAND_TOKENS}
    ca = {t for t in token_set(cand_addr) if len(t) >= 3 and t not in BRAND_TOKENS}
    if not ta or not ca:
        return 0.0
    inter = len(ta & ca)
    union = len(ta | ca)
    return 100.0 * (inter / union)

def haversine_km(lat1, lon1, lat2, lon2) -> float:
    try:
        lat1, lon1, lat2, lon2 = map(float, (lat1, lon1, lat2, lon2))
    except Exception:
        return 1e9
    R = 6371.0088
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (math.sin(dlat/2)**2 +
         math.cos(math.radians(lat1)) *
         math.cos(math.radians(lat2)) *
         math.sin(dlon/2)**2)
    return 2 * R * math.asin(math.sqrt(a))

def pin_equal(pin_input: str, pin_out1: str, addr_out1: str) -> bool:
    p1 = (pin_input or "").strip()
    p2 = (pin_out1 or "").strip()
    if p1 and p2 and p1 == p2:
        return True
    # fallback: check presence in candidate address
    return bool(p1) and (p1 in (addr_out1 or ""))

def distance_score_km(d_km: float) -> float:
    # 0–100, decays after ~50m, negligible after ~1.5km
    if d_km <= 0.05:
        return 100.0
    if d_km >= 1.5:
        return 0.0
    return max(0.0, 100.0 * (1.5 - d_km) / 1.5)

def build_match_score(
    assoc: str, tagged_addr: str, pin: str, olat: float, olong: float,
    cand_name: str, cand_addr: str, cand_pin: str, cand_lat: float, cand_lon: float,
    category: str, reviews: str
) -> Tuple[float, dict]:
    d_km = haversine_km(olat, olong, cand_lat, cand_lon)
    ds = distance_score_km(d_km)
    pin_ok = pin_equal(pin, cand_pin, cand_addr)
    bh = brand_hits(assoc, cand_name, cand_addr)
    loc = locality_hits(tagged_addr, cand_addr)
    cat_ok = (category or "").strip().lower() in WHITELIST_CATEGORIES

    try:
        rv = int(str(reviews or "0").replace(",", ""))
    except Exception:
        rv = 0
    rev_bonus = 3.0 if rv >= 1000 else (1.0 if rv >= 200 else 0.0)

    score = (
        0.45 * ds +
        0.25 * (100.0 if pin_ok else 0.0) +
        0.15 * min(100.0, bh * 50.0) +
        0.12 * loc +
        0.03 * (100.0 if cat_ok else 0.0) +
        rev_bonus
    )
    info = {"d_km": round(d_km, 4), "pin_ok": pin_ok, "brand_hits": bh, "loc_score": round(loc, 1), "rev": rv}
    return score, info

def hard_override(
    assoc: str, pin: str, olat: float, olong: float,
    cand_name: str, cand_addr: str, cand_pin: str, cand_lat: float, cand_lon: float
) -> bool:
    # Brand + near + pin → accept
    if brand_hits(assoc, cand_name, cand_addr) > 0:
        if haversine_km(olat, olong, cand_lat, cand_lon) <= 0.8 and pin_equal(pin, cand_pin, cand_addr):
            return True
    return False

# ================= MAIN =================
def main():
    # Prepare outputs
    ensure_excel_with_sheet(OUTPUT1_FILE, OUTPUT1_COLS, "Sheet1")
    ensure_excel_with_sheet(OUTPUT1_FILE, OUTPUT1_COLS, "Eliminated")

    progress = load_progress()
    last_done = int(progress.get("last_processed_input_row", -1))

    # Read input
    try:
        headers, input_rows = read_input_rows_xlsx(INPUT_FILE)
    except Exception as e:
        print(f"[{now_iso()}] FATAL: cannot read input: {e}")
        log_error(progress, -1, f"Input read error: {e}")
        return
    if not input_rows:
        print(f"[{now_iso()}] Empty input.")
        return

    # Validate required columns
    missing = [c for c in INPUT_REQUIRED if c not in headers]
    if missing:
        print(f"[{now_iso()}] FATAL: Missing columns in input: {missing}")
        return

    # Output 2 columns = input headers + OUTPUT1_COLS + FuzzyScore
    out2_cols = headers + OUTPUT1_COLS + ["FuzzyScore"]
    ensure_excel_with_sheet(OUTPUT2_FILE, out2_cols, "Sheet1")

    dedupe_db = open_dedupe_db()
    cache_db = open_cache_db()

    n = len(input_rows)
    print(f"[{now_iso()}] Loaded input with {n} rows. Resuming from row {last_done + 1}.")

    for i in range(last_done + 1, n):
        row = input_rows[i]
        try:
            assoc = row.get("Associate Hospital", "").strip()
            olat_raw = row.get("OLat", "")
            olong_raw = row.get("OLong", "")
            pin = row.get("pin", "")
            tagged = row.get("Tagged Address", "")

            olat = float_or_none(olat_raw)
            olong = float_or_none(olong_raw)
            if olat is None or olong is None:
                log_error(progress, i, f"Invalid coordinates: OLat='{olat_raw}', OLong='{olong_raw}'")
                progress["last_processed_input_row"] = i
                save_progress(progress)
                continue

            # purpose print
            print_call_purpose(i, olat, olong, pin, assoc, using_cache=False)

            js = call_serper_or_cache(cache_db, assoc, olat, olong, pin)
            if js is None:
                log_error(progress, i, "API call failed after retries")
                progress["last_processed_input_row"] = i
                save_progress(progress)
                continue

            places = js.get("places", []) or []

            # Collect candidates; per-row dedupe by stable key
            seen_keys = set()
            candidates: List[Dict[str, Any]] = []
            eliminated: List[Dict[str, Any]] = []

            for p in places[:NUM_RESULTS]:
                k = stable_place_key(p)
                if k in seen_keys:
                    continue
                seen_keys.add(k)

                name = p.get("title", "")
                addr = p.get("address", "")
                lat2 = p.get("latitude", "")
                lon2 = p.get("longitude", "")
                website = p.get("website", "") or ""
                cid = (p.get("cid") or "").strip()
                category = (p.get("category", "") or p.get("type", "") or "").strip()
                rating = p.get("rating", "")
                reviews = p.get("ratingCount", "")

                # Country/PIN guard
                if not in_india_or_has_pin(addr, lat2, lon2):
                    continue

                pinc = extract_pincode(addr)
                rec = {
                    "Name": name, "Address": addr, "Lat": lat2, "Long": lon2,
                    "Website": website, "Pincode": pinc, "Category": category,
                    "CID": cid, "InputRowIndex": i, "Rating": rating, "Reviews": reviews
                }

                # Whitelist classification (kept for Output1 vs Eliminated)
                catnorm = (category or "").lower().strip()
                if catnorm in WHITELIST_CATEGORIES:
                    candidates.append(rec)
                else:
                    eliminated.append(rec)

            # Funnel: keep nearest shortlist for scoring
            def dist_of(r):
                return haversine_km(olat, olong, r["Lat"] or 0, r["Long"] or 0)

            candidates.sort(key=dist_of)
            shortlist = candidates[:SHORTLIST_NEAREST]

            # Score shortlist; decide best
            best = None
            best_score = -1.0
            best_info = {}
            for r2 in shortlist:
                sc, info = build_match_score(
                    assoc, tagged, pin, olat, olong,
                    r2["Name"], r2["Address"], r2["Pincode"], r2["Lat"] or 0, r2["Long"] or 0,
                    r2["Category"], r2["Reviews"]
                )
                if hard_override(assoc, pin, olat, olong, r2["Name"], r2["Address"], r2["Pincode"], r2["Lat"] or 0, r2["Long"] or 0):
                    best, best_score, best_info = r2, 1000.0, {"override": True, **info}
                    break
                if sc > best_score:
                    best, best_score, best_info = r2, sc, info

            # Output 1 writing (allowed only) with global dedupe
            out1_to_write: List[Dict[str, Any]] = []
            for r2 in candidates:
                key2 = f"cid:{r2['CID']}" if r2.get("CID") else f"na:{norm_text(r2['Name'])}|{norm_text(r2['Address'])}"
                if is_new_and_mark(dedupe_db, key2):
                    out1_to_write.append(r2)
            if out1_to_write:
                append_rows_excel(OUTPUT1_FILE, out1_to_write, OUTPUT1_COLS, "Sheet1")
            if eliminated:
                append_rows_excel(OUTPUT1_FILE, eliminated, OUTPUT1_COLS, "Eliminated")

            # Output 2 writing (ONE best, if score high enough)
            out2_rows: List[Dict[str, Any]] = []
            if best and (best_score >= OUT2_MIN_SCORE):
                left = {c: row.get(c, "") for c in headers}
                out2_rows.append({**left, **best, "FuzzyScore": round(best_score, 2)})

            if out2_rows:
                append_rows_excel(OUTPUT2_FILE, out2_rows, out2_cols, "Sheet1")

            print(f"[{now_iso()}] ✅ Row {i}: "
                  f"{len(out1_to_write)} allowed→Output1, "
                  f"{len(eliminated)} eliminated, "
                  f"{1 if out2_rows else 0} best→Output2 "
                  f"(score={round(best_score,2) if best else 'NA'}, info={best_info}).")

            # Persist progress
            progress["last_processed_input_row"] = i
            save_progress(progress)

        except Exception as e:
            print(f"[{now_iso()}] ❌ Row {i} error: {e}")
            log_error(progress, i, str(e))
            progress["last_processed_input_row"] = i
            save_progress(progress)
            continue

    print(f"[{now_iso()}] All done. Progress saved to {PROGRESS_LOG.name}.")

if __name__ == "__main__":
    main()
