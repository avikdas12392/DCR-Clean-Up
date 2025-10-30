#!/usr/bin/env python3
import warnings; warnings.filterwarnings("ignore", message="The global interpreter lock")

import json
import time
import re
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple

import requests
from openpyxl import load_workbook, Workbook

# ---------- Optional fuzzy lib; fallback to difflib ----------
try:
    from rapidfuzz.fuzz import ratio as fuzz_ratio
    def fuzzy_score(a: str, b: str) -> float:
        return float(fuzz_ratio(a or "", b or ""))
except Exception:
    from difflib import SequenceMatcher
    def fuzzy_score(a: str, b: str) -> float:
        return 100.0 * SequenceMatcher(None, (a or ""), (b or "")).ratio()

# ================= CONFIG =================
API_KEY = "2e823b299673ed947dde2960a6b609da1f104d15"
ENDPOINT = "https://google.serper.dev/maps"

INPUT_FILE = Path("input.xlsx")
OUTPUT1_FILE = Path("output 1.xlsx")   # Sheet1 = allowed; Eliminated = filtered-out
OUTPUT2_FILE = Path("output 2.xlsx")
PROGRESS_LOG = Path("progress_log.json")

NUM_RESULTS = 20
RADIUS_METERS = 200
TIMEOUT = 30
BASE_SLEEP = 0.8
BACKOFF_FACTOR = 1.6
MAX_RETRIES = 4
FUZZY_THRESHOLD = 75.0
VICINITY_DECIMALS = 3

INDIA_LAT_MIN, INDIA_LAT_MAX = 6.0, 37.5
INDIA_LON_MIN, INDIA_LON_MAX = 68.0, 97.5

DEDUP_DB_FILE = Path("serper_seen.sqlite3")
CACHE_DB_FILE = Path("vicinity_cache.sqlite3")

OUTPUT1_COLS_FIXED = [
    "Name","Address","Lat","Long","Website","Pincode","Category","CID",
    "InputRowIndex","Rating","Reviews"
]

# ✅ Category whitelist (case-insensitive)
WHITELIST_CATEGORIES = {
    "hospital","nursing home","private hospital","general hospital",
    "government hospital","specialized hospital","heart hospital",
    "hospital department"
}

# ================= Utilities =================
def now_iso(): return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def load_progress():
    if not PROGRESS_LOG.exists():
        return {"last_processed_input_row": -1, "errors": []}
    try:
        return json.loads(PROGRESS_LOG.read_text(encoding="utf-8"))
    except Exception as e:
        PROGRESS_LOG.rename(PROGRESS_LOG.with_suffix(".corrupt.json"))
        return {"last_processed_input_row": -1,
                "errors":[{"row":-1,"error":f"Corrupt progress:{e}","time":now_iso()}]}

def save_progress(p): 
    tmp = PROGRESS_LOG.with_suffix(".tmp")
    tmp.write_text(json.dumps(p, indent=2), encoding="utf-8"); tmp.replace(PROGRESS_LOG)

def log_error(p, row, err):
    p.setdefault("errors", []).append({"row":row,"error":err,"time":now_iso()}); save_progress(p)

def print_call_purpose(i,olat,olon,pin,assoc,using_cache=False):
    src = "Cache" if using_cache else "Serper Maps API"
    print(f"[{now_iso()}] Row {i}: Search '{assoc or 'hospital'} + hospital' via {src}. "
          f"OLat={olat}, OLong={olon}, pin={pin}")

def extract_pincode(addr:str)->str:
    if not addr: return ""
    m = re.search(r"\b(\d{6})\b", addr)
    if m: return m.group(1)
    parts=[p.strip() for p in addr.split(",") if p.strip()]
    if len(parts)>=2:
        m2=re.search(r"\b(\d{6})\b",parts[-2])
        if m2:return m2.group(1)
    return ""

def stable_place_key(p):
    cid=(p.get("cid") or "").strip()
    if cid: return f"cid:{cid}"
    return f"na:{(p.get('title','').lower().strip())}|{(p.get('address','').lower().strip())}"

def float_or_none(x):
    try: return float(x)
    except: return None

def looks_like_india(lat,lon):
    try: return INDIA_LAT_MIN<=float(lat)<=INDIA_LAT_MAX and INDIA_LON_MIN<=float(lon)<=INDIA_LON_MAX
    except: return False

def in_india_or_has_pin(addr,lat,lon):
    try:
        if looks_like_india(lat,lon): return True
    except: pass
    a=(addr or "").lower()
    return " india" in a or bool(re.search(r"\b\d{6}\b",a))

# ---- Excel I/O ----
def ensure_excel_with_sheet(path,cols,sheet="Sheet1"):
    if path.exists():
        wb=load_workbook(path)
        if sheet in wb.sheetnames:
            ws=wb[sheet]
            if ws.max_row==0: ws.append(cols); wb.save(path)
            return
        ws=wb.create_sheet(title=sheet); ws.append(cols); wb.save(path); return
    wb=Workbook(); ws=wb.active; ws.title=sheet; ws.append(cols); wb.save(path)

def read_input_rows_xlsx(path):
    wb=load_workbook(path,data_only=True,read_only=True); ws=wb.active
    rows=list(ws.iter_rows(values_only=True))
    if not rows: return [],[]
    headers=[str(h or "").strip() for h in rows[0]]
    out=[{headers[j]:(str(r[j]) if r and j<len(r) and r[j] is not None else "")
          for j in range(len(headers))} for r in rows[1:]]
    return headers,out

def append_rows_excel(path,rows,cols,sheet="Sheet1"):
    if not rows: return
    ensure_excel_with_sheet(path,cols,sheet)
    wb=load_workbook(path); ws=wb[sheet]
    for r in rows: ws.append([r.get(c,"") for c in cols])
    wb.save(path)

# ---- SQLite helpers ----
def open_db(file,create_sql):
    db=sqlite3.connect(file); db.execute("PRAGMA journal_mode=WAL;"); db.execute("PRAGMA synchronous=NORMAL;")
    db.execute(create_sql); db.commit(); return db

def open_dedupe_db():
    return open_db(DEDUP_DB_FILE,
        "CREATE TABLE IF NOT EXISTS places_seen(key TEXT PRIMARY KEY, first_seen_at INTEGER DEFAULT (strftime('%s','now')));")

def is_new_and_mark(db,key):
    try: db.execute("INSERT INTO places_seen VALUES (?,strftime('%s','now'));",(key,)); db.commit(); return True
    except sqlite3.IntegrityError: return False

def open_cache_db():
    return open_db(CACHE_DB_FILE,
        "CREATE TABLE IF NOT EXISTS vicinity_cache(vkey TEXT PRIMARY KEY,response_json TEXT,last_used INTEGER DEFAULT (strftime('%s','now')));")

_in_mem_cache={}
def make_vkey(olat,olon,pin):
    def rd(x): 
        try:return f"{round(float(x),VICINITY_DECIMALS)}"
        except:return"NA"
    return f"{pin}|{rd(olat)}|{rd(olon)}"

def cache_get(db,vkey):
    js=_in_mem_cache.get(vkey)
    if js:return js
    cur=db.execute("SELECT response_json FROM vicinity_cache WHERE vkey=?",(vkey,))
    row=cur.fetchone()
    if row:
        try: js=json.loads(row[0]); _in_mem_cache[vkey]=js; return js
        except: return None
    return None

def cache_set(db,vkey,js):
    _in_mem_cache[vkey]=js
    db.execute("INSERT OR REPLACE INTO vicinity_cache VALUES(?, ?, strftime('%s','now'))",
               (vkey,json.dumps(js,separators=(',',':')))); db.commit()

# ---- Serper call ----
def call_serper_or_cache(cache_db,assoc,olat,olon,pin):
    vkey=make_vkey(olat,olon,pin)
    js=cache_get(cache_db,vkey)
    if js:
        print_call_purpose(-1,olat,olon,pin,assoc,True)
        print(f"[{now_iso()}]   ✅ Cache hit: {len(js.get('places',[]))} places."); return js

    headers={"X-API-KEY":API_KEY,"Content-Type":"application/json"}
    q=f"{(assoc or '').strip()} hospital near {olat},{olon} {pin} India".strip()
    payload={"q":q,"num":NUM_RESULTS,"start":0,"ll":f"{olat},{olon}","gl":"in","hl":"en"}
    if RADIUS_METERS>0: payload["radius"]=RADIUS_METERS

    for a in range(1,MAX_RETRIES+1):
        try:
            r=requests.post(ENDPOINT,headers=headers,json=payload,timeout=TIMEOUT)
            r.raise_for_status(); js=r.json()
            if "places" not in js: raise ValueError("No 'places'")
            print(f"[{now_iso()}]   ✅ API OK: {len(js['places'])} places."); cache_set(cache_db,vkey,js); return js
        except Exception as e:
            if a==MAX_RETRIES: print(f"[{now_iso()}] ❌ API failed: {e}"); return None
            t=BASE_SLEEP*(BACKOFF_FACTOR**(a-1))
            print(f"[{now_iso()}]   Retry {a}/{MAX_RETRIES} in {t:.1f}s: {e}"); time.sleep(t)
    return None

# ================= MAIN =================
def main():
    progress=load_progress(); last_done=int(progress.get("last_processed_input_row",-1))
    try: headers,rows=read_input_rows_xlsx(INPUT_FILE)
    except Exception as e: log_error(progress,-1,f"Input read:{e}"); return
    if not rows: print("Empty input."); return

    required=["Associate Hospital","OLat","OLong","pin","Tagged Address"]
    miss=[c for c in required if c not in headers]
    if miss: print(f"Missing columns:{miss}"); return

    out1_cols=OUTPUT1_COLS_FIXED[:]
    out2_cols=headers+out1_cols+["FuzzyScore"]

    ensure_excel_with_sheet(OUTPUT1_FILE,out1_cols,"Sheet1")
    ensure_excel_with_sheet(OUTPUT1_FILE,out1_cols,"Eliminated")
    ensure_excel_with_sheet(OUTPUT2_FILE,out2_cols)

    ddb=open_dedupe_db(); cdb=open_cache_db()
    print(f"[{now_iso()}] Loaded {len(rows)} rows. Resuming from {last_done+1}.")

    for i in range(last_done+1,len(rows)):
        r=rows[i]
        try:
            assoc=r.get("Associate Hospital","").strip(); olat=r.get("OLat",""); olon=r.get("OLong","")
            pin=r.get("pin",""); tagged=r.get("Tagged Address","")
            olatf=float_or_none(olat); olonf=float_or_none(olon)
            if olatf is None or olonf is None:
                log_error(progress,i,f"Invalid coords:{olat},{olon}"); progress["last_processed_input_row"]=i; save_progress(progress); continue

            print_call_purpose(i,olatf,olonf,pin,assoc)
            js=call_serper_or_cache(cdb,assoc,olatf,olonf,pin)
            if js is None: log_error(progress,i,"API fail"); progress["last_processed_input_row"]=i; save_progress(progress); continue
            places=js.get("places",[]) or []

            seen=set(); allowed=[]; eliminated=[]; out2=[]
            for p in places[:NUM_RESULTS]:
                key=stable_place_key(p)
                if key in seen: continue
                name,addr=p.get("title",""),p.get("address","")
                lat2,lon2=p.get("latitude",""),p.get("longitude","")
                website=p.get("website","") or ""; cid=(p.get("cid") or "").strip()
                category=(p.get("category","") or p.get("type","") or "").strip()
                rating,rev=p.get("rating",""),p.get("ratingCount","")

                if not in_india_or_has_pin(addr,lat2,lon2): continue
                pinc=extract_pincode(addr)
                rec={"Name":name,"Address":addr,"Lat":lat2,"Long":lon2,"Website":website,
                     "Pincode":pinc,"Category":category,"CID":cid,"InputRowIndex":i,
                     "Rating":rating,"Reviews":rev}

                seen.add(key)
                if not is_new_and_mark(ddb,key): continue

                catnorm=category.lower()
                if catnorm in WHITELIST_CATEGORIES:
                    allowed.append(rec)
                    # ✅ only whitelisted entries considered for Output 2 fuzzy match
                    sc=fuzzy_score(addr.lower(),tagged.lower())
                    if sc>=FUZZY_THRESHOLD:
                        left={c:r.get(c,"") for c in headers}
                        out2.append({**left,**rec,"FuzzyScore":round(sc,2)})
                else:
                    eliminated.append(rec)

            if allowed: append_rows_excel(OUTPUT1_FILE,allowed,out1_cols,"Sheet1")
            if eliminated: append_rows_excel(OUTPUT1_FILE,eliminated,out1_cols,"Eliminated")
            if out2: append_rows_excel(OUTPUT2_FILE,out2,out2_cols)
            print(f"[{now_iso()}] ✅ Row {i}: {len(allowed)} allowed, {len(eliminated)} eliminated, {len(out2)} fuzzy matches.")

            progress["last_processed_input_row"]=i; save_progress(progress)
        except Exception as e:
            log_error(progress,i,str(e)); progress["last_processed_input_row"]=i; save_progress(progress); continue

    print(f"[{now_iso()}] Done. Progress in {PROGRESS_LOG.name}.")

if __name__=="__main__": main()
